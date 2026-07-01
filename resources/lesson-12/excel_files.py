# Lesson 12 - Excel Files with openpyxl
# Run this script with:  python3 excel_files.py
#
# First install openpyxl:
#   pip3 install openpyxl
#
# This script can also CREATE the sample student_scores.xlsx file
# from student_scores.csv (run create_sample_xlsx() first).

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill


def create_sample_xlsx():
    """Convert the CSV sample data into an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # Bold header row with yellow background
    with open("student_scores.csv", newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            ws.append(row)
            if row_idx == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    wb.save("student_scores.xlsx")
    print("student_scores.xlsx created.")


def read_and_summarise():
    """Read the Excel file and print summary statistics."""
    wb = openpyxl.load_workbook("student_scores.xlsx")
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    print("Columns:", header)

    scores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        student_id, name, topic, score, grade = row
        if score is not None:
            scores.append(score)

    print(f"Students:      {len(scores)}")
    print(f"Average score: {sum(scores) / len(scores):.1f}")
    print(f"Highest score: {max(scores)}")
    print(f"Lowest score:  {min(scores)}")


def add_pass_fail_column():
    """Read student_scores.xlsx and save results.xlsx with a Pass/Fail column."""
    wb_in = openpyxl.load_workbook("student_scores.xlsx")
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"

    header = [cell.value for cell in ws_in[1]] + ["Pass/Fail"]
    ws_out.append(header)

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        student_id, name, topic, score, grade = row
        result = "Pass" if (score is not None and score >= 50) else "Fail"
        ws_out.append([student_id, name, topic, score, grade, result])

    wb_out.save("results.xlsx")
    print("results.xlsx saved.")


if __name__ == "__main__":
    # Run all steps
    create_sample_xlsx()
    read_and_summarise()
    add_pass_fail_column()
